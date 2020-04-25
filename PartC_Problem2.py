#!/usr/bin/env python
# coding: utf-8

# In[1]:


def update_attendance(excel_path, list_values, course_name, col_number):
    
    import pandas as pd
    import openpyxl
    import os
    
    data = pd.read_excel(excel_path, sheet_name = course_name)
    
    start_empty_value = []
    list_columns = []
    
    cols = data.columns
    
    for i in range(len(data.columns)):
        a = cols[i]
        if data[a].isnull().sum() != 0:
            list_columns.append(i)
            value = data[a].isnull().sum()
            empty_row = len(data) - value
            start_empty_value.append(empty_row)
                  
    wb = openpyxl.load_workbook(filename = excel_path)
    print(wb)
    
    head,tail = os.path.split(excel_path)
    
    ws = wb.sheetnames
    print(ws)

    for i in range(len(ws)):
        if ws[i] == course_name:
            new_ws = wb.worksheets[i]
            
    print(new_ws)
    
    new_predicted_values = []
    
    for j in range(len(list_values)):
        if list_values[j] == 0:
            new_predicted_values.append("A")
        else:
            new_predicted_values.append("P")
            
    print(new_predicted_values)
    
    print(start_empty_value)
            
    for i in range(0,len(new_predicted_values)):
        new_ws.cell(row = i + start_empty_value[col_number - 2] + 2, column = col_number).value = new_predicted_values[i]
        
    os.chdir(head)

    wb.save("Attendence_GoogleSheet.xlsx")
    
    return 'Successfully added columns to excel sheet'


# In[7]:


path = "E:\\Praxis\\Capstone Project\\29-03-2020\\Attendence_GoogleSheet.xlsx"
values = [0,1,1,1,1,1,0,0,1,0,1,1,1,1,0]
name = 'Python'
number = 9
update_attendance(path,values,name,number)


# In[ ]:


def update_attendance(excel_path, list_values, course_name, col_number):
    
    import openpyxl
    import os
    new_predicted_values = []
       
    wb = openpyxl.load_workbook(filename = excel_path)
    print(wb)
    
    head,tail = os.path.split(excel_path)
    
    ws = wb.sheetnames
    print(ws)

    for i in range(len(ws)):
        if ws[i] == course_name:
            new_ws = wb.worksheets[i]
            
    print(new_ws)
    
    for j in range(len(list_values)):
        if list_values[j] == 0:
            new_predicted_values.append("A")
        else:
            new_predicted_values.append("P")
            
    print(new_predicted_values)
            
    for i in range(0,len(new_predicted_values)):
        new_ws.cell(row=i+2, column = col_number).value = new_predicted_values[i]
        
    os.chdir(head)

    wb.save("attendence.xlsx")
    
    return 'Successfully added columns to excel sheet'